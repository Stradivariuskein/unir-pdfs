import xml.etree.ElementTree as ET
import tempfile
import zipfile
import os

import openpyxl


def get_xml_from_xlsx(ruta_xlsx, nombre_archivo_xml): # obtine un archivo xml de un arechivo excel
    # Crear un directorio temporal
    with tempfile.TemporaryDirectory() as directorio_temporal:
        # Descomprimir el archivo XLSX en el directorio temporal
        with zipfile.ZipFile(ruta_xlsx, 'r') as archivo_zip:
            archivo_zip.extractall(directorio_temporal)
        
        # Leer el contenido del archivo XML desde el directorio temporal
        ruta_xml = os.path.join(directorio_temporal, nombre_archivo_xml)
        try:
            with open(ruta_xml, 'r', encoding='utf-8') as archivo_xml:
                contenido_xml = archivo_xml.read()
            return contenido_xml
        except FileNotFoundError:
            print(f"El archivo XML '{nombre_archivo_xml}' no se encontró dentro del XLSX.")
            return None


def get_columns_from_xml(contenido_xml):
    # Crear un diccionario para almacenar las columnas y sus anchos
    cols = {}

    # Analizar el contenido XML si no está vacío
    if contenido_xml.strip():
        root = ET.fromstring(contenido_xml)

        # Buscar todas las etiquetas <col> dentro de <cols>
        ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        for col_elem in root.findall(".//ns:col", namespaces=ns):
            # Obtener el atributo "width" de la etiqueta <col>
            col_width = col_elem.attrib.get("width")

            # Si hay un atributo "width", usarlo para establecer el ancho de la columna
            if col_width:
                # Obtener el atributo "min" de la etiqueta <col>
                col_min = int(col_elem.attrib.get("min"))
                col_max = int(col_elem.attrib.get("max"))
                if col_min == col_max:
                    col_id = f"col{col_min}"
                    cols[col_id] = float(col_width)
                elif col_min < col_max:
                    while True:
                        if col_min > col_max:
                            break
                        col_id = f"col{col_min}"
                        cols[col_id] = float(col_width)
                        col_min += 1


    return cols

def get_imgs(hoja_excel): # devuelve una lista con todas las imagenes de un archivo
    lista = []
    for img in hoja_excel._images:
        lista.append(img)
        fila, columna = img.anchor._from.row, img.anchor._from.col
        columna_letra = openpyxl.utils.get_column_letter(columna+1)
        colOff, rowOff = img.anchor._from.colOff, img.anchor._from.rowOff
        print(fila, columna_letra)
        print(f"whidth: {img.width}, height: {img.height}")
        print(f"{colOff,rowOff}\n")
    return lista

def insert_img(img, cell, hoja_excel):
    
    # Coordenadas de celda y desplazamiento

    row, col = cell

    # guarda las propiedades de la imagen
    anchor_img = img.anchor

    # Insertar la imagen en la hoja de Excel
    hoja_excel.add_image(img, f'{col}{row}')
    img.anchor = anchor_img

    #img.anchor._from.colOff = colOff
    #img.anchor._from.rowOff = rowOff




if __name__ == "__main__":
    
    
    pxl_doc = openpyxl.load_workbook('ARANDELA CHAPISTA y comunes.xlsx')
    sheet = pxl_doc['Hoja1']


    images = get_imgs(sheet)

    wb_bacio = openpyxl.Workbook()
    hoja_actual = wb_bacio.active

    xml = get_xml_from_xlsx("ARANDELA CHAPISTA y comunes.xlsx", "xl/worksheets/sheet1.xml")
    columns = get_columns_from_xml(xml)


    hoja_actual.column_dimensions['A'].width = columns["col1"]
    hoja_actual.column_dimensions['B'].width = columns["col2"]
    hoja_actual.column_dimensions['C'].width = columns["col3"]
    hoja_actual.column_dimensions['D'].width = columns["col4"]
    hoja_actual.column_dimensions['E'].width = columns["col5"]

    insert_img(images[0],(1,1),hoja_actual)

    # Guardamos el archivo
    wb_bacio.save('nuevo_archivo.xlsx')