import openpyxl

import os
import fnmatch

def obtener_rutas_xlsx(directorio):
    lista_rutas_xlsx = []

    for ruta, _, archivos in os.walk(directorio):
        for archivo in archivos:
            if fnmatch.fnmatch(archivo, '*.xlsx'):
                ruta_completa = os.path.join(ruta, archivo)
                lista_rutas_xlsx.append(ruta_completa)

    return lista_rutas_xlsx





wb_bacio = openpyxl.Workbook()
#hoja_actual = wb_bacio.active
list_excels = obtener_rutas_xlsx("./LISTAS NUEVAS")
for excel in list_excels:

    pxl_doc = openpyxl.load_workbook(excel)
    sheet = pxl_doc['Hoja1']
    hoja_actual = wb_bacio.copy_worksheet(sheet)

    # Copiar contenidos de sheet a hoja_actual
    #for row in sheet.iter_rows(values_only=True):
        #hoja_actual.append(row)

wb_bacio.save("Catalogo.xlsx")