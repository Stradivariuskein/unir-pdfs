#Importing the modules
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.drawing.image import Image
from openpyxl import Workbook, drawing


import os
import zipfile
import xml.etree.ElementTree as ET
import tempfile

def obtener_xml_desde_xlsx(ruta_xlsx, nombre_archivo_xml):
    # Crear un directorio temporal
    with tempfile.TemporaryDirectory() as directorio_temporal:
        # Descomprimir el archivo XLSX en el directorio temporal
        with zipfile.ZipFile(ruta_xlsx, 'r') as archivo_zip:
            archivo_zip.extractall(directorio_temporal)
        
        # Leer el contenido del archivo XML desde el directorio temporal
        ruta_xml = os.path.join(directorio_temporal, nombre_archivo_xml)
        try:
            with open(ruta_xml, 'r', encoding='utf-8') as archivo_xml:
                contenido_xml = archivo_xml.read()
            return contenido_xml
        except FileNotFoundError:
            print(f"El archivo XML '{nombre_archivo_xml}' no se encontró dentro del XLSX.")
            return None


def obtener_columnas_desde_xml(contenido_xml):
    # Crear un diccionario para almacenar las columnas y sus anchos
    cols = {}

    # Analizar el contenido XML si no está vacío
    if contenido_xml.strip():
        root = ET.fromstring(contenido_xml)

        # Buscar todas las etiquetas <col> dentro de <cols>
        ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        for col_elem in root.findall(".//ns:col", namespaces=ns):
            # Obtener el atributo "width" de la etiqueta <col>
            col_width = col_elem.attrib.get("width")

            # Si hay un atributo "width", usarlo para establecer el ancho de la columna
            if col_width:
                # Obtener el atributo "min" de la etiqueta <col>
                col_min = int(col_elem.attrib.get("min"))
                col_max = int(col_elem.attrib.get("max"))
                if col_min == col_max:
                    col_id = f"col{col_min}"
                    cols[col_id] = float(col_width)
                elif col_min < col_max:
                    while True:
                        if col_min > col_max:
                            break
                        col_id = f"col{col_min}"
                        cols[col_id] = float(col_width)
                        col_min += 1


    return cols

def mostrar_valores_celdas(hoja_excel):
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    row = 1
    col = 0
    print(hoja_excel.column_dimensions['A'].width)
    print(hoja_excel.column_dimensions['B'].width)
    print(hoja_excel.column_dimensions["C"].width)
    print(hoja_excel.column_dimensions['D'].width)
    print(hoja_excel.column_dimensions['E'].width)
    for fila in hoja_excel.iter_rows():
        for celda in fila:
            if col > 10:
                col = 0
                break
            if celda.value != None:
                print(f"{cols[col]}{row}: {celda.value}")
            col += 1
        row += 1

def mostrar_imagenes(hoja_excel):
    lista = []
    for img in hoja_excel._images:
        lista.append(img)
        fila, columna = img.anchor._from.row, img.anchor._from.col
        columna_letra = openpyxl.utils.get_column_letter(columna+1)
        colOff, rowOff = img.anchor._from.colOff, img.anchor._from.rowOff
        print(fila, columna_letra)
        print(f"whidth: {img.width}, height: {img.height}")
        print(f"{colOff,rowOff}\n")
    return lista

def insertar_imagen(img, cell, ancho, alto, desplazamiento, hoja_excel):
    
    #img.width = ancho
    #img.height = alto

    # Coordenadas de celda y desplazamiento
    colOff, rowOff = desplazamiento
    row, col = cell

    # Ajustar el anclaje de la imagen
    anchor_img = img.anchor


    # Insertar la imagen en la hoja de Excel
    hoja_excel.add_image(img, f'{col}{row}')
    img.anchor = anchor_img
    img.width = ancho + 2000
    img.height = alto
    #img.anchor._from.colOff = colOff
    #img.anchor._from.rowOff = rowOff

xml = obtener_xml_desde_xlsx("ARANDELA CHAPISTA y comunes.xlsx", "xl/worksheets/sheet1.xml")
print(xml)
data = obtener_columnas_desde_xml(xml)
print(data)

wb_bacio = Workbook()
hoja_actual = wb_bacio.active

hoja_actual.column_dimensions['A'].width = 14
hoja_actual.column_dimensions['B'].width = 16.42578125
hoja_actual.column_dimensions['C'].width = 13
hoja_actual.column_dimensions['D'].width = 10.85546875
hoja_actual.column_dimensions['E'].width = 13.7109375

pxl_doc = openpyxl.load_workbook('ARANDELA CHAPISTA y comunes.xlsx')
sheet = pxl_doc['Hoja1']

ruta_imagen = 'logo.png'  # Reemplaza esto con la ruta de tu imagen
ancho_imagen = 385
alto_imagen = 177
desplazamiento_imagen = (952500, 38100)
cell = (1,2)
imagenes = mostrar_imagenes(sheet)
insertar_imagen(imagenes[0], cell, ancho_imagen, alto_imagen, desplazamiento_imagen, hoja_actual)

# Guardamos el archivo
wb_bacio.save('nuevo_archivo.xlsx')

#loading the Excel File and the sheet


mostrar_imagenes(sheet)
mostrar_valores_celdas(sheet)

#calling the image_loader
image_loader = SheetImageLoader(sheet)

#get the image (put the cell you need instead of 'A1')
image = image_loader.get('A1')

#showing the image
image.show()

#saving the image
image.save('arandelas.jpg')