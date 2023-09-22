import PyPDF2
import os

import openpyxl
import win32api
import win32print
import subprocess
import pyautogui
import time

def eliminar_paginas_en_blanco(archivo_entrada):
    pdf_reader = PyPDF2.PdfReader(archivo_entrada)
    pdf_writer = PyPDF2.PdfWriter()

    for num_pagina in range(len(pdf_reader.pages)):
        pagina = pdf_reader.pages[num_pagina]
        contenido = pagina.extract_text()

        # Verificar si el contenido de la página está vacío o solo contiene espacios en blanco
        if not contenido.strip():
            continue

        pdf_writer.add_page(pagina)


    pdf_writer.write(archivo_entrada)


def combinar_archivos_pdf(lista_archivos, archivo_salida):
    pdf_mezclado = PyPDF2.PdfWriter()

    for archivo in lista_archivos:
        eliminar_paginas_en_blanco(archivo)
        pdf_reader = PyPDF2.PdfReader(archivo)

        for num_pagina in range(len(pdf_reader.pages)):
            pagina = pdf_reader.pages[num_pagina]
            pdf_mezclado.add_page(pagina)

    with open(archivo_salida, 'wb') as archivo_salida:
        pdf_mezclado.write(archivo_salida)

def obtener_archivos_pdf(directorio):
    lista_archivos_pdf = []

    for nombre_archivo in os.listdir(directorio):
        ruta_archivo = os.path.join(directorio, nombre_archivo)

        if os.path.isfile(ruta_archivo) and nombre_archivo.lower().endswith(".pdf"):
            lista_archivos_pdf.append(ruta_archivo)

    return lista_archivos_pdf

from openpyxl import Workbook

def buscar_celda_con_palabra(hoja, palabra_buscada):
    try:

        # Buscar la palabra en todas las celdas de la hoja
        for fila in hoja.iter_rows():
            for celda in fila:
                if palabra_buscada in str(celda.value).lower():
                    return celda.coordinate

        # Si la palabra no se encuentra, devolver None
        return None

    except Exception as e:
        print(f"Error: {e}")
        return None
              


def borrar_celdas_desde(celda_inicio, hoja):
    try:
        # Obtener la columna y fila de la celda de inicio
        columna_inicio, fila_inicio = celda_inicio[0] , celda_inicio[1:]

        # Obtener la cantidad de filas en la hoja
        ultima_fila = hoja.max_row

        # Iterar sobre las filas, comenzando desde la fila de inicio, para borrar los valores
        for fila in range(int(fila_inicio), ultima_fila + 1):
            hoja[columna_inicio + str(fila)] = None
            hoja[columna_inicio + str(fila)].border = openpyxl.styles.Border()
            hoja[columna_inicio + str(fila)].fill = openpyxl.styles.PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")


        print(f"Valores borrados desde la celda {celda_inicio} hasta el final de la hoja.")

    except Exception as e:
        print(f"Error: {e}")

def listar_archivos_directorio(directorio): # obtiene de cada lista la ruta completa del drive
    ruta_listas = {}
    for raiz, directorios, archivos in os.walk(directorio):
        for archivo in archivos:
            if archivo[-5:] == ".xlsx": # si es un archivo excel
                ruta_xlsx = os.path.join(raiz, archivo)
                end_index = ruta_xlsx[::-1].find("\\") # quitamos el nombre de la lista de la ruta
                end_index = len(ruta_xlsx) - end_index
                ruta_listas[archivo] = ruta_xlsx[:end_index]
    return ruta_listas


def print_excel_to_pdf(file_paths, output_folder):
    for file_path in file_paths:
        # Comprobamos si la ruta es un archivo Excel (.xlsx)
        if file_path.lower().endswith('.xlsx'):
            try:
                # Cargamos el archivo Excel con openpyxl
                workbook = openpyxl.load_workbook(filename=file_path)
                # Obtenemos el nombre del archivo sin la extensión
                base_name = os.path.splitext(os.path.basename(file_path))[0]
                # Creamos el archivo PDF temporal para imprimirlo
                pdf_file_path = os.path.join(output_folder, f'{base_name}.pdf')

                # Impresión del archivo Excel a PDF usando la impresora predeterminada en segundo plano
                print_command = f'rundll32 printui.dll,PrintUIEntry /y /n "Microsoft Print to PDF" /t "{file_path}"'

                # Ejecutamos el comando
                subprocess.run(print_command, shell=True)

                print(f"Se ha iniciado la impresión de {file_path}. El PDF se guardará en {pdf_file_path}")

            except Exception as e:
                print(f"Error al procesar el archivo '{file_path}': {e}")
        else:
            print(f"El archivo '{file_path}' no es un archivo Excel (.xlsx)")




def quitar_precios(directorio):

    archivos_xlsx = listar_archivos_directorio(directorio)
    for name, rute in archivos_xlsx.items():
        libro = openpyxl.load_workbook(rute)
        hoja = libro['Hoja1']
        hoja["A1"] = None
        hoja["A1"].border = openpyxl.styles.Border()
        hoja["A1"].fill = openpyxl.styles.PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
        cell_precio = buscar_celda_con_palabra(hoja, "precio")
        hoja = borrar_celdas_desde(cell_precio,hoja)

        libro.save("./list_sin_precios/" + name)


def automate_save_pdf(ruta_archivos_excel):
    # Carpeta por defecto para guardar los archivos PDF
    output_folder = 'C:/Users/notebook/Desktop/catalogo/pdf'

    # Llamamos a la función para imprimir y guardar los archivos PDF
    print_excel_to_pdf(ruta_archivos_excel, output_folder)

    # Esperamos un tiempo para asegurarnos de que se haya iniciado la impresión
    time.sleep(2)

    # Esperamos un tiempo adicional para permitir que la impresión se complete (ajustar según la velocidad de impresión)
    time.sleep(10)


if __name__ == "__main__":


    automate_save_pdf(["C:/Users/notebook/Desktop/catalogo/list_sin_precios/BISAGRA T.xlsx"])
    #print_excel_to_pdf(["C:/Users/notebook/Desktop/catalogo/list_sin_precios/BISAGRA T.xlsx"])







# Ejemplo de uso:
#archivos_pdf = obtener_archivos_pdf("./pdfs")  # Reemplaza con la ruta de tu archivo PDF

#combinar_archivos_pdf(archivos_pdf,"Catalogo.pdf")